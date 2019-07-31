from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
from nltk.tokenize import word_tokenize, sent_tokenize
import openpyxl
import webbrowser
from gensim.summarization import keywords
from _operator import pos
from nltk.tag import pos_tag
from nltk.tag import untag 
from _ast import If
from flask import Flask, render_template
import flask


""" Creating a dictionary for storing the word frequency table."""
def generate_frequency_table(input_text) -> dict:
   
    """ Pre processing of data: Removing stop words"""
    stopWords = set(stopwords.words("english"))
    
    """Pre processing of data: Tokenization """
    words = word_tokenize(input_text)
    
    """ Pre processing of data: Stemming"""
    ps = PorterStemmer()
   

    frequencyTable = dict()
    for w in words:

        w = ps.stem(w)
        if w in stopWords:
            continue
        if w in frequencyTable:
            frequencyTable[w] += 1
        else:
            frequencyTable[w] = 1
    
    return frequencyTable

    

"""scoring sentences by its words
    Basic algorithm: adding the frequency of every non-stop word in a sentence divided by total no of words in a sentence."""
def scoring_input_sentences(input_sentences, freqTable) -> dict:
    

    sentenceScoreDict = dict()

    for s in input_sentences:
        word_count_in_sentence = (len(word_tokenize(s)))
        word_count_in_sentence_except_stop_words = 0
        for wordValue in freqTable:
            if wordValue in s.lower():
                word_count_in_sentence_except_stop_words += 1
                if s[:20] in sentenceScoreDict:
                    sentenceScoreDict[s[:20]] += freqTable[wordValue]
                else:
                    sentenceScoreDict[s[:20]] = freqTable[wordValue]

        sentenceScoreDict[s[:20]] = sentenceScoreDict[s[:20]] / word_count_in_sentence_except_stop_words
        
    return sentenceScoreDict
    

"""To calculate the average score"""
def calculate_average_score(sentenceScore) -> int:
    
    sumScores = 0
    for s in sentenceScore:
        sumScores += sentenceScore[s]

    avg_score = (sumScores / len(sentenceScore))
    
    return avg_score
    

"""To generate summary output"""
def create_summary(sentences, sentenceScore, threshold):
    sentence_counter = 0
    summary_output = ''

    for s in sentences:
        if s[:20] in sentenceScore and sentenceScore[s[:20]] > (threshold):
            summary_output += " " + s
            sentence_counter += 1

    return summary_output


def summarization(input_text):
    
    freq_table = generate_frequency_table(input_text)
    
    sentences = sent_tokenize(input_text)
   
    sentence_scores = scoring_input_sentences(sentences, freq_table)

    threshold_value = calculate_average_score(sentence_scores)
   
    summary_output = create_summary(sentences, sentence_scores, 0.92*threshold_value)
    
    return summary_output

def create_user_story(mlist):
    user_s=[]
    for m in mlist:
        m=m.lower()
        if m == "sound" or m == "sounds" or m=='music' or m=='physics' or m.startswith('environment') or m.startswith('interface') or m.startswith('movement') or m.startswith('control') or m=='audio' or m == "graphics" or m=='cutscenes' or m=='buttons':
            user_s.append(formatter(m,"better quality"))
        if m=='screen':
            user_s.append(formatter(m,"bigger size"))
        if m=='speed' or m.startswith('slow') or m.startswith('load') or m=='cpu' or m.startswith('lag'):
            user_s.append(formatter('games',"speedy and having good performance"))
        if m=='repetitive':
            user_s.append(formatter('games','non-repetitive'))
        if m=='fake' or m=='wrong' or m=='vague' or m=='confusing' or m=='boring' or m=='difficult':
            user_s.append(formatter('games','original and clear'))
        if m.startswith('defect') or m.startswith('install') or m=='incomplete' or m.startswith('finish') or m.startswith('issue') or m=='crappy' or m=='skip' or m.startswith('bug') or m=='fake' or m=='impossible' or m.startswith('crash') or m=='compatibility':
            user_s.append(formatter('games','defect free or bug free'))
        if m=='objective'or m.startswith('aim') or m.startswith('goal'):
            user_s.append(formatter('games','having good objectives or aims'))
        if m.startswith('level') or m.startswith('tracks') or m.startswith('cars'):
            user_s.append(formatter('games','having more '+m,))
        if m.startswith('collision'):
            user_s.append(formatter('games','collision free'))
        if m.startswith('glitch'):
            user_s.append(formatter('games','glitch free'))
    return user_s
    
def formatter(attribute1, attribute2):  
    
    a1=attribute1
    a2=attribute2
   
    user="As a user, I want {} to be {}.".format(a1,a2)
    
    return user

def final_user_story(flist):
    finallist=[]
    for i in flist:
        for j in i:
            finallist.append(j)
    #print(finallist)
    finallist=(list(dict.fromkeys(finallist)))
    #print(finallist)
    return finallist
    
def generate_html(finallist):
    app = Flask(__name__)

    @app.route("/")
    def home():
        return render_template('index.html', finallist=finallist)
    
    print("User stories generated from online user reviews")
    print("The number of user stories generated are: ",len(finallist))
    for i in finallist:
        print(i)
    print("Flask server starting")
    app.run(debug=True)
    
def read_and_write():
    
    flist=[]
    path = "C:\\Users\\gsree\\OneDrive\\Desktop\\Book2.xlsx"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row 
    for i in range(2, m_row+1): 
        cell_obj = sheet_obj.cell(row = i, column = 1)
        
        result = summarization(cell_obj.value)
        ci = sheet_obj.cell(row = i, column = 8)
        ci.value = result
        tok=word_tokenize(result)
        tagged=pos_tag(tok)
        nn_vb_tagged=[(word, tag) for word, tag in tagged if tag in ('NN', 'JJ','NNS','VBN','RB','VBG','VB')]
        mylist=untag(nn_vb_tagged)
        cj = sheet_obj.cell(row = i, column = 9)
        str=""
        for l in mylist:
            str=str+l+","
        cj.value=str
        mylist=(list(dict.fromkeys(mylist)))
        #print(mylist)
        ck = sheet_obj.cell(row = i, column = 10)
        string_story=[]
        string_story=(create_user_story(mylist))
        #print(string_story)
        
        st=""
        for i in string_story:
             st=st+i
        ck.value=st
        
        flist.append(string_story)
        #print(flist)
        
        
        wb_obj.save("C:\\Users\\gsree\\OneDrive\\Desktop\\Book2.xlsx") 
    
    generate_html(final_user_story(flist))
    
    
    
if __name__ == '__main__':
    print("Application started!!!")
    read_and_write()
    
    
    